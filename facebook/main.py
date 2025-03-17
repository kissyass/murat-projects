import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from openpyxl import load_workbook
import time, random, os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd

# -------------------------------------------------------------------
#                   CONFIGURATIONS & GLOBALS
# -------------------------------------------------------------------

CREDENTIALS_FILE = "fb_credentials.txt"
excel_file_path = None   # Path of the uploaded Excel file
processed_rows = []      # Final processed rows list
CURRENT_LANG = "en"      # Default language ("en" or "tr")

# 1) UI text in English & Turkish
UI_STRINGS = {
    "langLabel": {"en": "Language:", "tr": "Dil:"},
    "emailLabel": {"en": "Facebook Email/Phone:", "tr": "Facebook E-posta/Telefon:"},
    "passwordLabel": {"en": "Password:", "tr": "Şifre:"},
    "rememberMe": {"en": "Remember me", "tr": "Beni hatırla"},
    "noFile": {"en": "No file uploaded.", "tr": "Dosya yüklenmedi."},
    "uploadBtn": {"en": "Upload Excel File", "tr": "Excel Dosyası Yükle"},
    "processBtn": {"en": "Process Excel File", "tr": "Excel Dosyasını İşle"},
    "processing": {"en": "Processing data, please wait...", "tr": "Veriler işleniyor, lütfen bekleyin..."},
    "backBtn": {"en": "Back", "tr": "Geri"},
    "downloadBtn": {"en": "Download Output Excel", "tr": "Çıktıyı Excel Olarak İndir"},
    "successExport": {"en": "Data exported to", "tr": "Veriler şu konuma aktarıldı:"},
    "errNoData": {"en": "No data to export!", "tr": "Dışa aktarılacak veri yok!"},
    "errSaveFile": {"en": "Could not save file:\n", "tr": "Dosya kaydedilemedi:\n"},
    "errNoFile": {"en": "No file selected.", "tr": "Dosya seçilmedi."},
    "errNoCred": {"en": "Please enter Facebook credentials.", "tr": "Lütfen Facebook bilgilerinizi girin."},
    "errReadExcel": {"en": "Could not read the Excel file:\n", "tr": "Excel dosyası okunamadı:\n"},
    "errSelenium": {"en": "Error during Selenium processing:\n", "tr": "Selenium işlenmesi sırasında hata:\n"},

    # New strings for the "Done" message box
    "doneTitle": {"en": "Done", "tr": "Tamam"},
    "doneMessage": {"en": "Processing complete!", "tr": "İşlem tamamlandı!"}
}

# 2) Table column keys
TABLE_COLUMNS = ("group_link", "post_link", "post_status", "group_name", "group_type", "member_count")

# 3) Table headers in both languages
TABLE_HEADERS = {
    "group_link":   {"en": "Group Link",   "tr": "Grup Bağlantısı"},
    "post_link":    {"en": "Post Link",    "tr": "Gönderi Bağlantısı"},
    "post_status":  {"en": "Post Status",  "tr": "Gönderi Durumu"},
    "group_name":   {"en": "Group Name",   "tr": "Grup Adı"},
    "group_type":   {"en": "Group Type",   "tr": "Grup Türü"},
    "member_count": {"en": "Member Count", "tr": "Üye Sayısı"}
}

# 4) Post status translations
POST_STATUS_MAP = {
    "Published":    {"en": "Published",    "tr": "Yayınlandı"},
    "Pending":      {"en": "Pending",      "tr": "Onay Bekliyor"},
    "No permission":{"en": "No permission","tr": "İzni yok"},
    "Error":        {"en": "Error",        "tr": "Hata"}
}

# 5) Group type translations
GROUP_TYPE_MAP = {
    "Public":   {"en": "Public",   "tr": "Herkese Açık"},
    "Private":  {"en": "Private",  "tr": "Özel"},
    "Unknown":  {"en": "Unknown",  "tr": "Bilinmiyor"},
    "Not found":{"en": "Not found","tr": "Bulunamadı"}
}

# 6) For scraping fallback
NOT_FOUND_MAP = {
    "en": "Not found",
    "tr": "Bulunamadı"
}

# -------------------------------------------------------------------
#                 CREDENTIALS (SAVE/LOAD)
# -------------------------------------------------------------------

def load_credentials():
    try:
        with open(CREDENTIALS_FILE, "r", encoding="utf-8") as f:
            email = f.readline().strip()
            password = f.readline().strip()
        return email, password
    except Exception:
        return "", ""

def save_credentials(email, password):
    try:
        with open(CREDENTIALS_FILE, "w", encoding="utf-8") as f:
            f.write(email + "\n")
            f.write(password + "\n")
    except Exception as e:
        messagebox.showerror("Error", f"Could not save credentials:\n{e}")

# -------------------------------------------------------------------
#                       LANGUAGE HANDLING
# -------------------------------------------------------------------

def set_language(lang):
    """
    Update CURRENT_LANG and refresh all UI text (labels, buttons, table headings).
    Also update the table if it already has data displayed.
    """
    global CURRENT_LANG
    CURRENT_LANG = lang

    # Update labels & buttons
    lang_label.config(text=UI_STRINGS["langLabel"][lang])
    email_label.config(text=UI_STRINGS["emailLabel"][lang])
    password_label.config(text=UI_STRINGS["passwordLabel"][lang])
    remember_chk.config(text=UI_STRINGS["rememberMe"][lang])
    file_label.config(text=UI_STRINGS["noFile"][lang] if excel_file_path is None else f"{UI_STRINGS['noFile'][lang]}: {excel_file_path}")
    upload_btn.config(text=UI_STRINGS["uploadBtn"][lang])
    process_btn.config(text=UI_STRINGS["processBtn"][lang])
    back_btn.config(text=UI_STRINGS["backBtn"][lang])
    download_btn.config(text=UI_STRINGS["downloadBtn"][lang])

    # Update table headings
    for col in TABLE_COLUMNS:
        results_table.heading(col, text=TABLE_HEADERS[col][lang])

    # If we already have processed rows, re-display them with the new language
    if processed_rows:
        show_results_frame(processed_rows, refresh_only=True)

def localize_post_status(status):
    """Translate post status based on CURRENT_LANG."""
    return POST_STATUS_MAP.get(status, {"en": status, "tr": status})[CURRENT_LANG]

def localize_group_type(gtype):
    """Translate group type based on CURRENT_LANG."""
    return GROUP_TYPE_MAP.get(gtype, {"en": gtype, "tr": gtype})[CURRENT_LANG]

def localize_not_found():
    """Return the localized 'Not found' string."""
    return NOT_FOUND_MAP[CURRENT_LANG]

def localize_member_count(count_str):
    """
    If the language is Turkish, replace:
      'K' -> ' Bin'
      'M' -> ' Milyon'
      also replace 'members' -> 'üyeler', if desired
    """
    if CURRENT_LANG == "tr":
        # Quick replacements for thousand & million
        count_str = count_str.replace("K", " Bin")
        count_str = count_str.replace("M", " Milyon")
        count_str = count_str.replace("members", "üyeler")
    return count_str

# -------------------------------------------------------------------
#                   HELPER: CENTER THE WINDOW
# -------------------------------------------------------------------

def center_window(root, width=900, height=700):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coord = int((screen_width / 2) - (width / 2))
    y_coord = int((screen_height / 2) - (height / 2))
    root.geometry(f"{width}x{height}+{x_coord}+{y_coord}")

# -------------------------------------------------------------------
#               STRING PROCESSING / POST STATUS
# -------------------------------------------------------------------

def extract_group_link(text):
    """Extract the base group link from a post URL."""
    match = re.search(r"groups/(\d+)", text)
    if match:
        group_id = match.group(1)
        return f"https://www.facebook.com/groups/{group_id}/"
    return ""

def determine_post_status(displayed_text):
    """Determine post status based on displayed text."""
    if "permalink" in displayed_text:
        return "Published"
    elif "pending_posts" in displayed_text:
        return "Pending"
    elif "You don't have permission" in displayed_text:
        return "No permission"
    else:
        return "Error"

# -------------------------------------------------------------------
#                   FILE UPLOAD
# -------------------------------------------------------------------

def upload_file():
    global excel_file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm *.xlsb *.ods")],
        title=UI_STRINGS["uploadBtn"][CURRENT_LANG]
    )
    if file_path:
        excel_file_path = file_path
        file_label.config(text=f"{file_path}")
        process_btn.config(state="normal")

# -------------------------------------------------------------------
#                   SCRAPE GROUP INFO
# -------------------------------------------------------------------

def scrape_group_info(group_link, driver):
    """
    Given a group link and an active Selenium driver,
    navigate to the group page and scrape the group's name, type, and member count.
    Returns a tuple: (group_name, group_type, member_count)
    """
    driver.get(group_link)
    time.sleep(5)  # Wait for page to load

    try:
        group_name = driver.find_element(By.XPATH, "//h1").text
    except Exception:
        group_name = "Not found"

    group_type = "Unknown"
    try:
        # Check if it's a public group
        privacy_elem = driver.find_element(By.XPATH, "//*[contains(translate(text(), 'PUBLIC', 'public'), 'public group')]")
        if privacy_elem:
            group_type = "Public"
    except:
        try:
            privacy_elem = driver.find_element(By.XPATH, "//*[contains(translate(text(), 'PRIVATE', 'private'), 'private group')]")
            if privacy_elem:
                group_type = "Private"
        except Exception:
            group_type = "Unknown"

    try:
        members_elem = driver.find_element(By.XPATH, "//*[contains(translate(text(), 'MEMBERS', 'members'), 'members')]")
        member_count = members_elem.text
    except Exception:
        member_count = "Not found"

    return group_name, group_type, member_count

# -------------------------------------------------------------------
#                 MAIN PROCESSING FUNCTION
# -------------------------------------------------------------------

def process_file():
    global processed_rows

    if not excel_file_path:
        messagebox.showerror("Error", UI_STRINGS["errNoFile"][CURRENT_LANG])
        return

    fb_email = email_entry.get().strip()
    fb_password = password_entry.get().strip()
    if not fb_email or not fb_password:
        messagebox.showerror("Error", UI_STRINGS["errNoCred"][CURRENT_LANG])
        return

    # Save credentials if "Remember me" is checked
    if remember_var.get():
        save_credentials(fb_email, fb_password)

    status_label.config(text=UI_STRINGS["processing"][CURRENT_LANG])
    progress_bar["value"] = 0
    root.update_idletasks()

    # Read the Excel file
    try:
        wb = load_workbook(filename=excel_file_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"{UI_STRINGS['errReadExcel'][CURRENT_LANG]}{e}")
        status_label.config(text="")
        return

    rows_data = []
    for row in sheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        displayed_text = str(cell.value).strip() if cell.value else ""
        effective_link = cell.hyperlink.target if cell.hyperlink else displayed_text

        post_status = determine_post_status(displayed_text)
        group_link = extract_group_link(effective_link)

        rows_data.append({
            "group_link": group_link,
            "post_link": effective_link,
            "post_status": post_status,
            "group_name": "",
            "group_type": "",
            "member_count": ""
        })

    unique_groups = {r["group_link"] for r in rows_data if r["group_link"]}
    group_info = {}

    # Setup Selenium (headless, English)
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--lang=en")
    # If Chrome is not in default location, specify it like:
    # chrome_options.binary_location = r"C:\Path\to\chrome.exe"

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get('https://www.facebook.com/')
        time.sleep(3)
        driver.find_element(By.ID, 'email').send_keys(fb_email)
        driver.find_element(By.ID, 'pass').send_keys(fb_password)
        driver.find_element(By.NAME, 'login').click()
        time.sleep(5)

        total_groups = len(unique_groups)
        progress_bar["maximum"] = total_groups
        current = 0

        for grp in unique_groups:
            if grp:
                info = scrape_group_info(grp, driver)
                group_info[grp] = info
                current += 1
                progress_bar["value"] = current
                root.update_idletasks()
                time.sleep(random.uniform(5, 15))
    except Exception as e:
        messagebox.showerror("Error", f"{UI_STRINGS['errSelenium'][CURRENT_LANG]}{e}")
    finally:
        driver.quit()

    # Merge the scraped info
    for r in rows_data:
        grp = r["group_link"]
        if grp in group_info:
            g_name, g_type, m_count = group_info[grp]
            r["group_name"] = g_name
            r["group_type"] = g_type
            r["member_count"] = m_count
        else:
            r["group_name"] = localize_not_found()
            r["group_type"] = "Unknown"
            r["member_count"] = localize_not_found()

    # Sort rows by post status
    order = {"Published": 1, "Pending": 2, "No permission": 3, "Error": 4}
    rows_data.sort(key=lambda x: order.get(x["post_status"], 99))
    processed_rows = rows_data

    # Localize final data if the user is in Turkish
    localize_rows_data(processed_rows)

    # Show results in the UI
    show_results_frame(processed_rows)
    status_label.config(text="")

    # ---- SHOW A "DONE" MESSAGE BOX ----
    messagebox.showinfo(UI_STRINGS["doneTitle"][CURRENT_LANG], UI_STRINGS["doneMessage"][CURRENT_LANG])

def localize_rows_data(rows_data):
    """
    Convert post_status, group_type, and member_count to Turkish if CURRENT_LANG is 'tr'.
    E.g., K->Bin, M->Milyon, 'Public'->'Herkese Açık', etc.
    """
    for row in rows_data:
        # Post status
        row["post_status"] = localize_post_status(row["post_status"])
        # Group type
        row["group_type"] = localize_group_type(row["group_type"])
        # Member count
        row["member_count"] = localize_member_count(row["member_count"])
        # If group_name was "Not found", localize it
        if row["group_name"] == NOT_FOUND_MAP["en"] or row["group_name"] == NOT_FOUND_MAP["tr"]:
            row["group_name"] = localize_not_found()

# -------------------------------------------------------------------
#         SHOWING RESULTS & REFRESHING TABLE HEADINGS
# -------------------------------------------------------------------

def show_results_frame(rows_data, refresh_only=False):
    """
    Display rows_data in the results_table.
    If refresh_only=True, it means we're only updating the table headings and re-inserting the same data
    in a new language (the underlying data has already been localized).
    """
    for i in results_table.get_children():
        results_table.delete(i)

    for row in rows_data:
        results_table.insert("", "end", values=(
            row["group_link"],
            row["post_link"],
            row["post_status"],
            row["group_name"],
            row["group_type"],
            row["member_count"]
        ))

    if not refresh_only:
        main_frame.pack_forget()
        results_frame.pack(fill="both", expand=True)

def show_main_frame():
    results_frame.pack_forget()
    main_frame.pack(fill="both", expand=True)

# -------------------------------------------------------------------
#             DOWNLOAD (EXPORT) TO EXCEL
# -------------------------------------------------------------------

def download_excel():
    if not processed_rows:
        messagebox.showerror("Error", UI_STRINGS["errNoData"][CURRENT_LANG])
        return
    df = pd.DataFrame(processed_rows)
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx *.xls")],
        title="Save as Excel"
    )
    if save_path:
        try:
            df.to_excel(save_path, index=False)
            messagebox.showinfo("Success", f"{UI_STRINGS['successExport'][CURRENT_LANG]} {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"{UI_STRINGS['errSaveFile'][CURRENT_LANG]}{e}")

# -------------------------------------------------------------------
#                         GUI SETUP
# -------------------------------------------------------------------

root = tk.Tk()
root.title("Facebook Group & Post Processor")
center_window(root, width=900, height=700)

# Frames
main_frame = tk.Frame(root)
results_frame = tk.Frame(root)

# We use grid in results_frame so we can fix ~2/3 table height
results_frame.grid_rowconfigure(0, weight=2)  # 2/3
results_frame.grid_rowconfigure(1, weight=1)  # 1/3
results_frame.grid_columnconfigure(0, weight=1)

# ----------------- Language Selection -------------------
lang_frame = tk.Frame(main_frame)
lang_label = tk.Label(lang_frame, text="Language:")  # Will be updated by set_language
language_var = tk.StringVar(value="en")

def on_language_select(event=None):
    chosen_lang = language_var.get()
    set_language(chosen_lang)

lang_combo = ttk.Combobox(lang_frame, textvariable=language_var, values=["en", "tr"], width=5, state="readonly")
lang_combo.bind("<<ComboboxSelected>>", on_language_select)

lang_label.pack(side="left", padx=5)
lang_combo.pack(side="left")
lang_frame.pack(pady=5)

# ----------------- Credentials Frame -------------------
credentials_frame = tk.Frame(main_frame)
email_label = tk.Label(credentials_frame, text="Facebook Email/Phone:")
password_label = tk.Label(credentials_frame, text="Password:")
email_entry = tk.Entry(credentials_frame, width=30)
password_entry = tk.Entry(credentials_frame, width=30, show="*")
remember_var = tk.BooleanVar()
remember_chk = tk.Checkbutton(credentials_frame, text="Remember me", variable=remember_var)

email_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)
email_entry.grid(row=0, column=1, padx=5, pady=5)
password_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
password_entry.grid(row=1, column=1, padx=5, pady=5)
remember_chk.grid(row=2, column=1, sticky="w", pady=5)

# Load saved credentials if available
saved_email, saved_password = load_credentials()
if saved_email and saved_password:
    email_entry.insert(0, saved_email)
    password_entry.insert(0, saved_password)

credentials_frame.pack(pady=10)

# ----------------- File upload & Process -------------------
file_label = tk.Label(main_frame, text="No file uploaded.")
upload_btn = tk.Button(main_frame, text="Upload Excel File", command=upload_file)
process_btn = tk.Button(main_frame, text="Process Excel File", state="disabled", command=process_file)
status_label = tk.Label(main_frame, text="", fg="blue")
progress_bar = ttk.Progressbar(main_frame, orient="horizontal", length=300, mode="determinate")

file_label.pack(pady=5)
upload_btn.pack(pady=5)
process_btn.pack(pady=5)
status_label.pack(pady=5)
progress_bar.pack(pady=5)

main_frame.pack(fill="both", expand=True)

# ----------------- Results Frame (Table + Buttons) -------------------
# The table in row=0 => 2/3 of the frame. Buttons in row=1 => 1/3
table_frame = tk.Frame(results_frame)
table_frame.grid(row=0, column=0, sticky="nsew")

columns = TABLE_COLUMNS
results_table = ttk.Treeview(table_frame, columns=columns, show="headings")
for col in columns:
    results_table.heading(col, text=TABLE_HEADERS[col]["en"])  # Default in English
    results_table.column(col, width=150, anchor="center")

scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=results_table.yview)
results_table.configure(yscroll=scrollbar.set)
results_table.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

buttons_frame = tk.Frame(results_frame)
buttons_frame.grid(row=1, column=0, sticky="nsew")

download_btn = tk.Button(buttons_frame, text="Download Output Excel", command=download_excel)
back_btn = tk.Button(buttons_frame, text="Back", command=show_main_frame)
download_btn.pack(pady=5)
back_btn.pack(pady=5)

# ----------------- Initialize UI in English, but user can switch  -------------------
on_language_select()  # triggers set_language("en") by default

root.mainloop()
